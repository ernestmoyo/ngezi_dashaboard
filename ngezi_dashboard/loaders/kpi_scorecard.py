"""
Loader for KPIs - Q3 FY2020 scorecard.

The source file has a two-row merged header (rows 3-4, 1-indexed).
Row 3 carries period groups; row 4 carries sub-columns (Actual, Budget, Var %).
Data runs rows 5-18 with KPI labels in column B.
"""

import logging

import openpyxl
import pandas as pd

from .utils import safe_float, normalise_percentage

logger = logging.getLogger(__name__)

# Expected period groups from row 3 (left to right)
_PERIOD_GROUPS = [
    ("q1", "Quarter 1"),
    ("q2_ytd", "Quarter 2 to date"),
    ("aug", "August"),
    ("q3", "Quarter 3"),
    ("fy20_ytd", "FY2020 YTD"),
]

# Sub-columns under each period group
_SUB_COLS = ["actual", "budget", "var_pct"]


def load_kpi_scorecard(path: str) -> pd.DataFrame:
    """Load FY20 Q3 KPI scorecard from Excel.

    Assumptions
    -----------
    - Header spans rows 3-4 (1-indexed).
    - Row 3: period group labels in merged cells.
    - Row 4: sub-column labels (Actual, Budget, Var %).
    - KPI labels are in column B (index 1).
    - First two data rows (Tis, BMS external audits) are non-plant KPIs
      and are dropped.
    - Data rows run from row 5 to row 18.
    - Comments are in column R (index 17, 0-based).

    Returns
    -------
    DataFrame with columns:
        kpi, q1_actual, q1_budget, q1_var_pct, ...,
        fy20_ytd_actual, fy20_ytd_budget, fy20_ytd_var_pct, comments
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        logger.exception("Failed to open KPI scorecard: %s", path)
        raise

    sheet_name = "FY20 Q3"
    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]
        logger.warning("Sheet 'FY20 Q3' not found, using '%s'", sheet_name)

    ws = wb[sheet_name]

    # Build column mapping from the merged header structure.
    # Row 3 has period groups starting at column C (index 3, 1-based).
    # Each group spans 3 columns (Actual, Budget, Var%).
    # Column B (col 2) = KPI label.
    # Column R (col 18) = Comments.

    # Map: column index (1-based) -> (period_prefix, sub_col)
    col_map: dict[int, str] = {}

    # Period groups start at column C (col 3).
    # Q1: C,D,E (3,4,5)  |  Q2 YTD: F,G,H (6,7,8)  |  Aug: I,J,K (9,10,11)
    # Q3: L,M,N (12,13,14)  |  FY20 YTD: O,P,Q (15,16,17)
    col_start = 3
    for prefix, _ in _PERIOD_GROUPS:
        for i, sub in enumerate(_SUB_COLS):
            col_map[col_start + i] = f"{prefix}_{sub}"
        col_start += 3

    # Non-plant KPI labels to skip
    skip_labels = {"Tis", "BMS external audits"}

    rows = []
    for row_idx in range(5, 19):  # rows 5-18 (1-indexed)
        kpi_label = ws.cell(row=row_idx, column=2).value
        if kpi_label is None:
            continue
        kpi_label = str(kpi_label).strip()
        if kpi_label in skip_labels or not kpi_label:
            continue

        record: dict = {"kpi": kpi_label}

        for col_idx, col_name in col_map.items():
            raw = ws.cell(row=row_idx, column=col_idx).value
            val = safe_float(raw)
            # Variance columns are already in percentage form in the source
            record[col_name] = val

        # Comments column (R = column 18)
        comments_raw = ws.cell(row=row_idx, column=18).value
        record["comments"] = str(comments_raw).strip() if comments_raw else None

        rows.append(record)

    wb.close()

    if not rows:
        logger.warning("No KPI rows extracted from %s", path)

    df = pd.DataFrame(rows)

    # Ensure numeric columns are float64
    numeric_cols = [c for c in df.columns if c not in ("kpi", "comments")]
    for col in numeric_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    logger.info("Loaded %d KPI rows from %s", len(df), path)
    return df
