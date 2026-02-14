"""
Loader for mill-ball procurement, consumption forecast, and grind projection.

Source: Copy_of_Mill_ball_trends_rev_1.xlsx

Structure per sheet (Sheet1, Sheet2, Changed):
    Rows 3-7: Supplier-level ball inventory (Vega, Maggotteaux, Anhui, Frusina)
    Row 10 onward: Monthly projected milled tonnage and consumption calculations
    Row 10 col D = "Projected Milled tonnage"; cols N-W = monthly dates
    Row 18-20: Mill 1 (Anhui) consumption/stock
    Row 25-29: Mill 2 (Vega/Maggoteaux/Frusina) consumption/stock

Dates may be datetime objects or Excel serial numbers (especially in 'Changed' sheet).
"""

import logging

import openpyxl
import pandas as pd

from .utils import normalise_date, safe_float

logger = logging.getLogger(__name__)

# Monthly date columns span N-W (cols 14-23, 1-indexed)
_DATE_COL_START = 14
_DATE_COL_END = 23


def _extract_monthly_series(
    ws, date_row: int, data_rows: dict[str, int], col_start: int, col_end: int
) -> list[dict]:
    """Extract a monthly time series from date headers + data rows.

    Parameters
    ----------
    ws : openpyxl worksheet
    date_row : 1-based row index containing month dates
    data_rows : mapping of field name -> 1-based row index
    col_start, col_end : 1-based column range (inclusive)

    Returns
    -------
    List of dicts, one per month, with 'month' and field columns.
    """
    records = []
    for col_idx in range(col_start, col_end + 1):
        raw_date = ws.cell(row=date_row, column=col_idx).value
        month = normalise_date(raw_date)
        if month is None:
            continue

        record = {"month": month}
        for field_name, row_idx in data_rows.items():
            record[field_name] = safe_float(ws.cell(row=row_idx, column=col_idx).value)
        records.append(record)

    return records


def load_mill_ball_trends(path: str, sheet_name: str = "Sheet1") -> pd.DataFrame:
    """Load mill-ball monthly consumption and stock projection.

    Assumptions
    -----------
    - Row 10 (cols N-W): monthly dates for the projection horizon.
    - Row 11: projected milled tonnage (with growth factor applied).
    - Row 18: consumption rate (g/t) for Mill 1 (Anhui).
    - Row 19: steel usage (tonnes) for Mill 1.
    - Row 20: remaining stock after Mill 1 consumption.
    - Row 25: consumption rate for Mill 2.
    - Row 26-27: steel usage for Mill 2 suppliers.

    Parameters
    ----------
    path : Path to the Excel file.
    sheet_name : Which sheet to load ('Sheet1', 'Sheet2', or 'Changed').

    Returns
    -------
    DataFrame with columns:
        month, projected_milled_tonnage, mill1_consumption_gt,
        mill1_steel_t, mill1_stock_remaining, mill2_consumption_gt
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        logger.exception("Failed to open mill ball file: %s", path)
        raise

    if sheet_name not in wb.sheetnames:
        logger.warning("Sheet '%s' not found, using '%s'", sheet_name, wb.sheetnames[0])
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    # Extract monthly time series
    data_rows = {
        "projected_milled_tonnage": 11,
        "mill1_consumption_gt": 18,
        "mill1_steel_t": 19,
        "mill1_stock_remaining": 20,
    }

    records = _extract_monthly_series(
        ws,
        date_row=10,
        data_rows=data_rows,
        col_start=_DATE_COL_START,
        col_end=_DATE_COL_END,
    )

    # Also try to get Mill 2 consumption from row 25
    for i, col_idx in enumerate(range(_DATE_COL_START, _DATE_COL_END + 1)):
        if i < len(records):
            val = safe_float(ws.cell(row=25, column=col_idx).value)
            records[i]["mill2_consumption_gt"] = val

    wb.close()

    df = pd.DataFrame(records)
    if not df.empty:
        df["month"] = pd.to_datetime(df["month"])
        numeric_cols = [c for c in df.columns if c != "month"]
        for col in numeric_cols:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    logger.info(
        "Loaded %d monthly mill-ball records from %s [%s]",
        len(df), path, sheet_name,
    )
    return df


def load_supplier_inventory(path: str, sheet_name: str = "Sheet1") -> pd.DataFrame:
    """Load supplier-level ball inventory from rows 3-7.

    Returns
    -------
    DataFrame with columns: supplier, ball_size, unit_weight_kg
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        logger.exception("Failed to open mill ball file: %s", path)
        raise

    if sheet_name not in wb.sheetnames:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    rows = []
    for row_idx in range(3, 8):
        supplier = ws.cell(row=row_idx, column=2).value
        if supplier is None:
            continue
        supplier = str(supplier).strip()
        if not supplier or supplier.lower() in ("stock", "mille"):
            continue

        ball_size = ws.cell(row=row_idx, column=3).value
        unit_weight = safe_float(ws.cell(row=row_idx, column=4).value)

        rows.append({
            "supplier": supplier,
            "ball_size": str(ball_size).strip() if ball_size else None,
            "unit_weight_kg": unit_weight,
        })

    wb.close()

    df = pd.DataFrame(rows)
    logger.info("Loaded %d supplier inventory rows from %s", len(df), path)
    return df
